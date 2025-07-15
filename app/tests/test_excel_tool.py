"""
Tests for the Excel Export Tool
"""

import os
import shutil
import tempfile
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.tools.excel_tool import ExcelExportTool


class TestExcelExportTool:
    """Tests for the Excel Export Tool"""

    @pytest.fixture(scope="class")
    def excel_tool(self):
        """Create and return an instance of ExcelExportTool"""
        tool = ExcelExportTool()
        yield tool

    @pytest.fixture(scope="function")
    def test_dir(self):
        """Create a temporary test directory for output files"""
        original_dir = os.getcwd()
        # Create temp directory
        temp_dir = tempfile.mkdtemp()
        os.chdir(temp_dir)

        # Create training_plans directory inside the temp dir
        os.makedirs(os.path.join(temp_dir, "training_plans"), exist_ok=True)

        yield temp_dir

        # Cleanup
        os.chdir(original_dir)
        shutil.rmtree(temp_dir)

    def test_basic_export(self, excel_tool, test_dir):
        """Test basic export functionality"""
        # Setup test data
        athlete_name = "Test Runner"
        week_start_date = "2025-07-15"

        workouts = [
            {
                "day": "Monday",
                "title": "Easy Run",
                "duration": "45 min",
                "description": "Easy recovery run at conversational pace",
                "type": "Run",
                "intensity": "Easy",
            },
            {
                "day": "Wednesday",
                "title": "Interval Training",
                "duration": "60 min",
                "description": "6 x 400m repeats with 200m recovery jogs",
                "type": "Run",
                "intensity": "Hard",
            },
            {
                "day": "Friday",
                "title": "Long Run",
                "duration": "90 min",
                "description": "Steady long run to build endurance",
                "type": "Run",
                "intensity": "Moderate",
            },
        ]

        notes = "Focus on proper form and hydration this week."

        # Execute the tool
        result = excel_tool._run(
            athlete_name=athlete_name,
            week_start_date=week_start_date,
            workouts=workouts,
            notes=notes,
        )

        # Print result for debugging
        print(f"\nExcel Export Result:\n{result}\n")

        # Verify success message
        assert "exported successfully" in result

        # Extract file path from result
        file_path = result.split("exported successfully to ")[1].strip()

        # Verify file exists
        assert os.path.exists(file_path)

        # Load the Excel file and verify content
        wb = load_workbook(file_path)

        # Check that expected sheets exist
        assert "Overview" in wb.sheetnames
        assert "Weekly Plan" in wb.sheetnames
        assert "Notes" in wb.sheetnames

        # Check overview sheet
        overview = wb["Overview"]
        assert f"Training Plan for: {athlete_name}" in overview["A1"].value
        assert f"Week: {week_start_date}" in overview["A2"].value

        # Check workout data
        weekly_plan = wb["Weekly Plan"]
        assert weekly_plan["A2"].value == "Monday"  # First workout day
        assert weekly_plan["B2"].value == "Easy Run"  # First workout title

        # Check notes
        notes_sheet = wb["Notes"]
        assert notes_sheet["A2"].value == notes

    def test_different_workouts(self, excel_tool, test_dir):
        """Test exporting different types of workouts"""
        athlete_name = "Multi-Sport Athlete"
        week_start_date = "2025-07-15"

        # Mixed workout types
        workouts = [
            {
                "day": "Monday",
                "title": "Swim Technique",
                "duration": "45 min",
                "description": "Focus on stroke technique and efficiency",
                "type": "Swim",
                "intensity": "Moderate",
            },
            {
                "day": "Tuesday",
                "title": "Strength Session",
                "duration": "60 min",
                "description": "Full body strength workout with focus on core",
                "type": "Strength",
                "intensity": "Moderate",
            },
            {
                "day": "Thursday",
                "title": "Bike Intervals",
                "duration": "75 min",
                "description": "5 x 5-minute intervals at threshold power",
                "type": "Bike",
                "intensity": "Hard",
            },
            {
                "day": "Saturday",
                "title": "Yoga Recovery",
                "duration": "45 min",
                "description": "Restorative yoga session for recovery",
                "type": "Yoga",
                "intensity": "Easy",
            },
        ]

        # Execute the tool without notes
        result = excel_tool._run(
            athlete_name=athlete_name,
            week_start_date=week_start_date,
            workouts=workouts,
            notes=None,
        )

        # Verify success message
        assert "exported successfully" in result

        # Extract file path from result
        file_path = result.split("exported successfully to ")[1].strip()

        # Verify file exists
        assert os.path.exists(file_path)

        # Load and check the workbook
        wb = load_workbook(file_path)

        # Notes sheet shouldn't exist since no notes were provided
        assert "Notes" not in wb.sheetnames

        # Check workout data for different sport types
        weekly_plan = wb["Weekly Plan"]

        # Extract sport types from the sheet (column E should have the type)
        sport_types = []
        for row in range(2, 6):  # Rows 2-5 contain our 4 workouts
            sport_types.append(weekly_plan[f"E{row}"].value)

        # Verify all sport types are included
        assert "Swim" in sport_types
        assert "Strength" in sport_types
        assert "Bike" in sport_types
        assert "Yoga" in sport_types

    def test_error_handling(self, excel_tool, test_dir):
        """Test error handling with invalid inputs"""
        # Test with invalid date format
        result = excel_tool._run(
            athlete_name="Invalid Date Athlete",
            week_start_date="not-a-date",
            workouts=[
                {
                    "day": "Monday",
                    "title": "Test",
                    "duration": "30 min",
                    "description": "Test workout",
                    "type": "Run",
                    "intensity": "Easy",
                }
            ],
        )

        # Should return error message
        assert "Error creating Excel export" in result

        # Test with empty workouts list
        result = excel_tool._run(
            athlete_name="No Workouts Athlete",
            week_start_date="2025-07-15",
            workouts=[],
        )

        # Should still create a file, just with no workout data
        assert "exported successfully" in result

    def test_file_naming(self, excel_tool, test_dir):
        """Test file naming conventions"""
        # Test with spaces in athlete name
        athlete_name = "John Doe Smith"
        week_start_date = "2025-07-15"

        workouts = [
            {
                "day": "Monday",
                "title": "Test Workout",
                "duration": "30 min",
                "description": "Test description",
                "type": "Run",
                "intensity": "Easy",
            }
        ]

        result = excel_tool._run(
            athlete_name=athlete_name,
            week_start_date=week_start_date,
            workouts=workouts,
        )

        # Extract file path
        file_path = result.split("exported successfully to ")[1].strip()
        file_name = os.path.basename(file_path)

        # Check that spaces are replaced with underscores in the filename
        assert "John_Doe_Smith" in file_name
        assert week_start_date in file_name


if __name__ == "__main__":
    # For manual execution
    tool = ExcelExportTool()

    # Set up test data
    athlete = "Test Athlete"
    date = datetime.now().strftime("%Y-%m-%d")
    test_workouts = [
        {
            "day": "Monday",
            "title": "Test Run",
            "duration": "45 min",
            "description": "Test workout description",
            "type": "Run",
            "intensity": "Moderate",
        }
    ]

    # Run the tool
    result = tool._run(
        athlete_name=athlete, week_start_date=date, workouts=test_workouts
    )
    print(f"Test result: {result}")
